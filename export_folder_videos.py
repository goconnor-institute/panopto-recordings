#!/usr/bin/env python3
"""
Panopto Folder Video Exporter
------------------------------
Reads pt_class_groups.xlsx, fetches all videos from every IOE source folder
via the Panopto API, and writes an Excel workbook with:
  - A "Summary" sheet (one row per folder with video counts)
  - One sheet per folder containing all its videos

Usage:
    python export_folder_videos.py
    python export_folder_videos.py --output my_export.xlsx
    python export_folder_videos.py --only 201907   # single class group ID
    python export_folder_videos.py --include-bc    # also export BC folders
"""

import argparse
import base64
import json
import math
import os
import sys
from datetime import datetime, timedelta

import openpyxl
import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

load_dotenv(override=True)

# ── Configuration ──────────────────────────────────────────────────────────────
PANOPTO_SERVER   = os.getenv("PANOPTO_SERVER", "")
PANOPTO_CLIENT_ID     = os.getenv("PANOPTO_CLIENT_ID", "")
PANOPTO_CLIENT_SECRET = os.getenv("PANOPTO_CLIENT_SECRET", "")
TOKEN_FILE       = os.getenv("TOKEN_FILE", "panopto_tokens.json")
CLASS_GROUPS_FILE = os.getenv("CLASS_GROUPS_FILE", "pt_class_groups.xlsx")
PAGE_SIZE = 100          # Panopto API page size

# Header styling colours
COLOUR_HEADER_SUMMARY = "1F4E79"   # dark blue
COLOUR_HEADER_FOLDER  = "2F5496"   # medium blue
COLOUR_SUBHEADER      = "BDD7EE"   # light blue
COLOUR_ALT_ROW        = "EBF3FB"   # very light blue
DEFAULT_EXPORT_DIR = "folder-exports"

# ── Authentication ─────────────────────────────────────────────────────────────

def _load_tokens():
    if not os.path.exists(TOKEN_FILE):
        return None
    with open(TOKEN_FILE) as f:
        return json.load(f)

def _save_tokens(access_token, refresh_token, expires_in):
    data = {
        "access_token": access_token,
        "refresh_token": refresh_token,
        "expires_at": (datetime.now() + timedelta(seconds=expires_in - 300)).isoformat(),
        "created_at": datetime.now().isoformat(),
    }
    with open(TOKEN_FILE, "w") as f:
        json.dump(data, f, indent=2)

def _refresh(refresh_token):
    url = f"https://{PANOPTO_SERVER}/Panopto/oauth2/connect/token"
    creds = base64.b64encode(f"{PANOPTO_CLIENT_ID}:{PANOPTO_CLIENT_SECRET}".encode()).decode()
    resp = requests.post(
        url,
        data={"grant_type": "refresh_token", "refresh_token": refresh_token},
        headers={
            "Content-Type": "application/x-www-form-urlencoded",
            "Authorization": f"Basic {creds}",
        },
        timeout=30,
    )
    if resp.status_code == 200:
        td = resp.json()
        _save_tokens(td["access_token"], td.get("refresh_token", refresh_token), td.get("expires_in", 3600))
        return td["access_token"]
    print(f"❌ Token refresh failed: {resp.status_code} – {resp.text[:200]}")
    return None

def get_auth_token():
    """Return a valid Panopto access token, refreshing if needed.
    If the refresh token is expired, instructs the user to run panopto_reauth.py."""
    td = _load_tokens()
    if td:
        try:
            if datetime.now() < datetime.fromisoformat(td["expires_at"]):
                return td["access_token"]
        except Exception:
            pass
        token = _refresh(td.get("refresh_token", ""))
        if token:
            return token

    print("⚠️  Token refresh failed or no valid tokens found.")
    print("🔑  Run the following to re-authenticate, then retry:")
    print(f"      python panopto_reauth.py")
    sys.exit(1)

# ── Panopto API helpers ────────────────────────────────────────────────────────

def get_folder_info(token, folder_id):
    """Fetch folder metadata (name, etc.) from Panopto."""
    url = f"https://{PANOPTO_SERVER}/Panopto/api/v1/folders/{folder_id}"
    resp = requests.get(url, headers={"Authorization": f"Bearer {token}"}, timeout=30)
    if resp.status_code == 200:
        return resp.json()
    return None

def get_all_sessions(token, folder_id):
    """
    Page through all sessions in a folder and return a list of dicts.
    Handles pagination automatically.
    """
    sessions = []
    page = 0
    while True:
        url = f"https://{PANOPTO_SERVER}/Panopto/api/v1/folders/{folder_id}/sessions"
        params = {
            "sortField": "CreatedDate",
            "sortOrder": "Desc",
            "pageNumber": page,
            "pageSize": PAGE_SIZE,
        }
        resp = requests.get(
            url,
            headers={"Authorization": f"Bearer {token}"},
            params=params,
            timeout=30,
        )
        if resp.status_code == 403:
            print(f"    ⚠️  Access denied to folder {folder_id}")
            break
        if resp.status_code == 404:
            print(f"    ⚠️  Folder not found: {folder_id}")
            break
        if resp.status_code != 200:
            print(f"    ❌ API error {resp.status_code} for folder {folder_id}: {resp.text[:200]}")
            break

        data = resp.json()
        batch = data.get("Results", [])
        sessions.extend(batch)

        total = data.get("Total", 0)
        if len(sessions) >= total or not batch:
            break
        page += 1

    return sessions

def format_duration(seconds):
    """Convert a duration in seconds (float/int) to HH:MM:SS string."""
    if seconds is None:
        return ""
    try:
        s = int(float(seconds))
        h = s // 3600
        m = (s % 3600) // 60
        sec = s % 60
        return f"{h:02d}:{m:02d}:{sec:02d}"
    except (ValueError, TypeError):
        return str(seconds)

def format_date(iso_str):
    """Convert an ISO datetime string to a readable local format."""
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return str(iso_str)

def session_url(session_id):
    return f"https://{PANOPTO_SERVER}/Panopto/Pages/Viewer.aspx?id={session_id}"


# ── Excel helpers ──────────────────────────────────────────────────────────────

def _header_cell(ws, row, col, value, bg_hex, font_colour="FFFFFF", bold=True, font_size=11):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, color=font_colour, size=font_size)
    cell.fill = PatternFill("solid", fgColor=bg_hex)
    cell.alignment = Alignment(wrap_text=False, vertical="center")
    return cell

def _safe_sheet_name(name, max_len=31):
    """Sanitise a string for use as an Excel sheet name."""
    invalid = r'\/:*?"<>|'
    for ch in invalid:
        name = name.replace(ch, "_")
    name = name.strip()
    return name[:max_len] if len(name) > max_len else name

def _make_unique_name(name, existing):
    """Append a counter suffix to make sheet name unique."""
    if name not in existing:
        return name
    i = 2
    while f"{name[:28]}_{i}" in existing:
        i += 1
    return f"{name[:28]}_{i}"

def resolve_output_path(output_arg):
    """Resolve output path so bare filenames default to ./folder-exports."""
    output_arg = output_arg.strip()
    if not output_arg:
        output_arg = f"panopto_videos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    if os.path.dirname(output_arg):
        return output_arg

    return os.path.join(DEFAULT_EXPORT_DIR, output_arg)

# ── Core logic ─────────────────────────────────────────────────────────────────

def load_class_groups(include_bc=False):
    """
    Load pt_class_groups.xlsx and return a list of dicts with folder metadata.
    Each entry contains at minimum: class_group_id, shortname, folder_id, folder_type.
    """
    df = pd.read_excel(CLASS_GROUPS_FILE)

    entries = []
    for _, row in df.iterrows():
        raw_cg = row.get("Class Group ID", "")
        # Normalise float IDs (e.g. 201907.0 → "201907")
        if pd.notna(raw_cg):
            try:
                cg_id = str(int(float(raw_cg)))
            except (ValueError, TypeError):
                cg_id = str(raw_cg).strip()
        else:
            cg_id = ""
        shortname = row.get("Shortname", "") or ""

        ioe_id = row.get("IOE Folder ID")
        if pd.notna(ioe_id) and str(ioe_id).strip():
            entries.append({
                "class_group_id": cg_id,
                "shortname": str(shortname).strip(),
                "folder_id": str(ioe_id).strip(),
                "folder_type": "IOE",
            })

        if include_bc:
            bc_id = row.get("BC Folder ID")
            if pd.notna(bc_id) and str(bc_id).strip():
                entries.append({
                    "class_group_id": cg_id,
                    "shortname": str(shortname).strip(),
                    "folder_id": str(bc_id).strip(),
                    "folder_type": "BC",
                })

    return entries

def build_workbook(folder_results):
    """
    Build and return an openpyxl Workbook from the collected folder results.

    folder_results: list of dicts:
        {
          "class_group_id": str,
          "shortname": str,
          "folder_id": str,
          "folder_type": str,           # "IOE" or "BC"
          "folder_name": str,           # from Panopto API
          "sessions": [...],            # raw Panopto session dicts
          "error": str | None,
        }
    """
    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    sum_headers = [
        "Class Group ID", "Shortname", "Folder Type",
        "Panopto Folder Name", "Folder ID", "Video Count", "Notes",
    ]
    for col, h in enumerate(sum_headers, 1):
        _header_cell(ws_sum, 1, col, h, COLOUR_HEADER_SUMMARY, font_size=11)

    ws_sum.row_dimensions[1].height = 20
    ws_sum.freeze_panes = "A2"

    sum_row = 2
    used_sheet_names = {"Summary"}

    # ── One sheet per folder ───────────────────────────────────────────────────
    VIDEO_COLS = ["#", "Session Name", "Start Date/Time", "Duration", "Session ID", "URL"]
    COL_WIDTHS  = [5,    50,            18,                12,         38,           55]

    for result in folder_results:
        folder_name  = result.get("folder_name") or result["folder_id"]
        class_group  = result["class_group_id"]
        shortname    = result["shortname"]
        folder_type  = result["folder_type"]
        folder_id    = result["folder_id"]
        sessions     = result.get("sessions", [])
        error        = result.get("error")

        # ── Summary row ────────────────────────────────────────────────────────
        note = error if error else f"{len(sessions)} video(s)"
        row_data = [class_group, shortname, folder_type, folder_name, folder_id, len(sessions), note]
        for col, val in enumerate(row_data, 1):
            cell = ws_sum.cell(row=sum_row, column=col, value=val)
            if sum_row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=COLOUR_ALT_ROW)
        sum_row += 1

        # ── Individual folder sheet ────────────────────────────────────────────
        raw_name = f"{shortname} ({folder_type})" if shortname else folder_name
        sheet_name = _safe_sheet_name(raw_name)
        sheet_name = _make_unique_name(sheet_name, used_sheet_names)
        used_sheet_names.add(sheet_name)

        ws = wb.create_sheet(sheet_name)

        # Folder info header block (rows 1-2)
        ws.merge_cells("A1:F1")
        title_cell = ws["A1"]
        title_cell.value = f"{folder_name}  |  {folder_type} Folder  |  {shortname}  |  ID: {folder_id}"
        title_cell.font = Font(bold=True, color="FFFFFF", size=12)
        title_cell.fill = PatternFill("solid", fgColor=COLOUR_HEADER_FOLDER)
        title_cell.alignment = Alignment(wrap_text=False, vertical="center")
        ws.row_dimensions[1].height = 22

        # Column headers (row 2)
        for col, (h, w) in enumerate(zip(VIDEO_COLS, COL_WIDTHS), 1):
            _header_cell(ws, 2, col, h, COLOUR_SUBHEADER, font_colour="1F4E79", font_size=10)
            ws.column_dimensions[get_column_letter(col)].width = w

        ws.freeze_panes = "A3"

        if not sessions:
            ws.merge_cells("A3:F3")
            ws["A3"].value = error if error else "No sessions found in this folder."
            ws["A3"].font = Font(italic=True, color="888888")
        else:
            for i, session in enumerate(sessions, 1):
                data_row = 2 + i
                row_vals = [
                    i,
                    session.get("Name", ""),
                    format_date(session.get("StartTime") or session.get("CreatedDate")),
                    format_duration(session.get("Duration")),
                    session.get("Id", ""),
                    session_url(session.get("Id", "")),
                ]
                for col, val in enumerate(row_vals, 1):
                    cell = ws.cell(row=data_row, column=col, value=val)
                    cell.alignment = Alignment(vertical="top", wrap_text=False)
                    if i % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=COLOUR_ALT_ROW)

                # Make URL a hyperlink
                url_cell = ws.cell(row=data_row, column=6)
                url_cell.hyperlink = url_cell.value
                url_cell.font = Font(color="0563C1", underline="single",
                                     italic=False,
                                     bold=False)

    # ── Auto-fit summary columns ───────────────────────────────────────────────
    SUM_WIDTHS = [16, 30, 12, 40, 38, 12, 30]
    for col, w in enumerate(SUM_WIDTHS, 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = w

    return wb

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Export all videos from Panopto source folders to Excel"
    )
    parser.add_argument(
        "--output", "-o",
        default=os.path.join(
            DEFAULT_EXPORT_DIR,
            f"panopto_videos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        ),
        help="Output Excel filename (default: folder-exports/panopto_videos_export_<timestamp>.xlsx)",
    )
    parser.add_argument(
        "--only", "-f",
        default=None,
        help="Process only the folder matching this Class Group ID",
    )
    parser.add_argument(
        "--include-bc",
        action="store_true",
        help="Also export BC (destination) folders in addition to IOE folders",
    )
    args = parser.parse_args()
    args.output = resolve_output_path(args.output)

    print("=" * 65)
    print("🎬  PANOPTO FOLDER VIDEO EXPORTER")
    print("=" * 65)

    if not os.path.exists(CLASS_GROUPS_FILE):
        print(f"❌ Class groups file not found: {CLASS_GROUPS_FILE}")
        sys.exit(1)

    token = get_auth_token()

    entries = load_class_groups(include_bc=args.include_bc)

    if args.only:
        entries = [e for e in entries if str(e["class_group_id"]) == str(args.only)]
        if not entries:
            print(f"❌ No folder found with Class Group ID: {args.only}")
            sys.exit(1)
        print(f"🔍 Filtering to Class Group ID: {args.only}")

    total = len(entries)
    print(f"📋 {total} folder(s) to process from {CLASS_GROUPS_FILE}\n")

    folder_results = []

    for idx, entry in enumerate(entries, 1):
        folder_id   = entry["folder_id"]
        folder_type = entry["folder_type"]
        shortname   = entry["shortname"]
        cg_id       = entry["class_group_id"]

        label = f"{shortname or cg_id} [{folder_type}]"
        print(f"  [{idx:>3}/{total}]  {label:<35} {folder_id}", flush=True)

        # Fetch folder name from API
        folder_info = get_folder_info(token, folder_id)
        if folder_info:
            folder_name = folder_info.get("Name", folder_id)
        else:
            folder_name = folder_id

        # Fetch all sessions with pagination
        try:
            sessions = get_all_sessions(token, folder_id)
            error = None
        except requests.exceptions.Timeout:
            sessions = []
            error = "Timeout fetching sessions"
            print(f"         ⏰ Timeout")
        except Exception as exc:
            sessions = []
            error = str(exc)
            print(f"         ❌ Error: {exc}")

        video_count = len(sessions)
        if error is None:
            print(f"         ✅ {video_count} video(s)  —  \"{folder_name}\"")

        folder_results.append({
            **entry,
            "folder_name": folder_name,
            "sessions": sessions,
            "error": error,
        })

    print()
    print("📊  Building Excel workbook …", flush=True)
    wb = build_workbook(folder_results)

    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    try:
        wb.save(args.output)
    except PermissionError:
        print(f"❌ Could not save '{args.output}' because the file is in use.")
        print("   Close the workbook in Excel and run again, or use a different --output name.")
        sys.exit(1)
    total_videos = sum(len(r["sessions"]) for r in folder_results)
    print(f"✅  Saved: {args.output}")
    print(f"    {len(folder_results)} folder(s)  |  {total_videos} total video(s)")
    print("=" * 65)


if __name__ == "__main__":
    main()
