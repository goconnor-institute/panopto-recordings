#!/usr/bin/env python3
"""
Moodle Course ID Populator
--------------------------
Reads pt_class_groups.xlsx, looks up each course shortname via the
Moodle REST API, and writes the Moodle Course ID into the spreadsheet.

Usage:
    python populate_moodle_course_ids.py              # populate missing IDs only
    python populate_moodle_course_ids.py --force      # overwrite existing IDs too
    python populate_moodle_course_ids.py --dry-run    # preview without writing
"""

import os
import sys
import argparse
import time
from datetime import datetime

import requests
import openpyxl
from dotenv import load_dotenv

load_dotenv(override=True)

# Configuration
CLASS_GROUPS_FILE = os.getenv("CLASS_GROUPS_FILE", "pt_class_groups.xlsx")
MOODLE_URL = os.getenv("MOODLE_URL", "https://moodle.instituteofeducation.ie")
MOODLE_TOKEN = os.getenv("MOODLE_API_KEY", "")

# Column name to write into
COURSE_ID_COLUMN = "Moodle Course ID"
SHORTNAME_COLUMN = "Shortname"


def find_column_index(ws, col_name):
    """Find the 1-based column index for a column header name."""
    for cell in ws[1]:
        if cell.value and str(cell.value).strip() == col_name:
            return cell.column
    return None


def lookup_course_by_shortname(shortname):
    """Look up a Moodle course ID by its exact shortname via REST API."""
    try:
        r = requests.get(
            f"{MOODLE_URL}/webservice/rest/server.php",
            params={
                "wstoken": MOODLE_TOKEN,
                "wsfunction": "core_course_get_courses_by_field",
                "moodlewsrestformat": "json",
                "field": "shortname",
                "value": shortname,
            },
            timeout=30,
        )
        if r.status_code != 200:
            return None, f"HTTP {r.status_code}"

        data = r.json()

        # Check for API errors
        if "exception" in data:
            return None, data.get("message", "API error")

        courses = data.get("courses", [])
        if not courses:
            return None, "No course found"

        # Return the first match
        course = courses[0]
        return course["id"], course.get("fullname", "")

    except requests.RequestException as e:
        return None, str(e)


def main():
    parser = argparse.ArgumentParser(
        description="Populate Moodle Course IDs in pt_class_groups.xlsx"
    )
    parser.add_argument("--force", "-f", action="store_true",
                        help="Overwrite existing Course IDs")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview lookups without writing to the file")
    args = parser.parse_args()

    if not MOODLE_TOKEN:
        print("❌ MOODLE_API_KEY not set in .env")
        sys.exit(1)

    if not os.path.exists(CLASS_GROUPS_FILE):
        print(f"❌ File not found: {CLASS_GROUPS_FILE}")
        sys.exit(1)

    # Load workbook
    wb = openpyxl.load_workbook(CLASS_GROUPS_FILE)
    ws = wb.active

    # Find columns
    shortname_col = find_column_index(ws, SHORTNAME_COLUMN)
    course_id_col = find_column_index(ws, COURSE_ID_COLUMN)

    if not shortname_col:
        print(f"❌ Column '{SHORTNAME_COLUMN}' not found in spreadsheet")
        sys.exit(1)

    if not course_id_col:
        # Add the column
        course_id_col = ws.max_column + 1
        ws.cell(row=1, column=course_id_col, value=COURSE_ID_COLUMN)
        print(f"➕ Added '{COURSE_ID_COLUMN}' column at position {course_id_col}")

    total_rows = ws.max_row - 1  # Exclude header
    print("=" * 70)
    print("🎓 MOODLE COURSE ID POPULATOR")
    print(f"📋 File: {CLASS_GROUPS_FILE}")
    print(f"📊 Total rows: {total_rows}")
    if args.force:
        print("🔄 FORCE MODE: Will overwrite existing IDs")
    if args.dry_run:
        print("👁️ DRY RUN: No changes will be saved")
    print("=" * 70)

    # Process each row
    found = 0
    skipped = 0
    not_found = 0
    errors = 0
    updated = 0

    for row_num in range(2, ws.max_row + 1):
        shortname_cell = ws.cell(row=row_num, column=shortname_col)
        course_id_cell = ws.cell(row=row_num, column=course_id_col)

        shortname = shortname_cell.value
        if not shortname or not str(shortname).strip():
            continue

        shortname = str(shortname).strip()
        existing_id = course_id_cell.value

        # Skip if already has an ID (unless --force)
        if existing_id and str(existing_id).strip() and not args.force:
            skipped += 1
            continue

        # Look up course
        course_id, info = lookup_course_by_shortname(shortname)

        row_label = f"[{row_num - 1}/{total_rows}]"

        if course_id:
            found += 1
            if not args.dry_run:
                course_id_cell.value = course_id
                updated += 1
            status = "✅"
            print(f"  {row_label} {status} {shortname[:45]:<47} -> {course_id}  ({info[:40]})")
        elif "No course found" in str(info):
            not_found += 1
            status = "⚠️"
            print(f"  {row_label} {status} {shortname[:45]:<47} -> NOT FOUND")
        else:
            errors += 1
            status = "❌"
            print(f"  {row_label} {status} {shortname[:45]:<47} -> ERROR: {info}")

        # Small delay to avoid hammering the API
        time.sleep(0.2)

    # Save
    if not args.dry_run and updated > 0:
        backup_name = CLASS_GROUPS_FILE.replace(".xlsx", f"_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(backup_name)
        print(f"\n💾 Backup saved: {backup_name}")

        wb.save(CLASS_GROUPS_FILE)
        print(f"💾 Updated: {CLASS_GROUPS_FILE}")

    # Summary
    print(f"\n{'=' * 70}")
    print("📊 SUMMARY")
    print(f"  Found:      {found}")
    print(f"  Skipped:    {skipped} (already had IDs)")
    print(f"  Not found:  {not_found}")
    print(f"  Errors:     {errors}")
    if not args.dry_run:
        print(f"  Written:    {updated}")
    print("=" * 70)

    if errors > 0:
        sys.exit(1)


if __name__ == "__main__":
    main()
